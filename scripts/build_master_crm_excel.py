import json
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

output_xlsx_path = "/Volumes/homes/Kevin/AI_BUILDS/AKCONCERTS-COM_AG-v101/akconcerts-com/AK_Concerts_Master_CMS_CRM_Database.xlsx"
artifact_xlsx_path = "/Users/kb/.gemini/antigravity-ide/brain/68236e0c-08a9-4d0b-903b-7eb1a98165a7/AK_Concerts_Master_CMS_CRM_Database.xlsx"

events_path = "/Volumes/homes/Kevin/AI_BUILDS/AKCONCERTS-COM_AG-v101/akconcerts-com/src/data/events.json"
venues_path = "/Volumes/homes/Kevin/AI_BUILDS/AKCONCERTS-COM_AG-v101/akconcerts-com/venues.json"
bands_ts_path = "/Volumes/homes/Kevin/AI_BUILDS/AKCONCERTS-COM_AG-v101/akconcerts-com/src/data/bands.ts"

with open(events_path, 'r', encoding='utf-8') as f:
    events_data = json.load(f)

with open(venues_path, 'r', encoding='utf-8') as f:
    venues_data = json.load(f)

# Parse all bands from bands.ts
bands_data = []
with open(bands_ts_path, 'r', encoding='utf-8') as f:
    ts_content = f.read()
    
# Extract array content
match = re.search(r'export const bands: Band\[\] = \[\s*([\s\S]*?)\s*\];', ts_content)
if match:
    raw_array = match.group(1)
    # Parse individual band objects
    obj_matches = re.findall(r'\{\s*(.*?)\s*\}', raw_array)
    for obj_str in obj_matches:
        band = {}
        for prop in ['name', 'slug', 'description', 'facebookUrl', 'amazonUrl', 'youtubeId']:
            m_prop = re.search(rf'{prop}:\s*"(.*?)"', obj_str)
            if m_prop:
                band[prop] = m_prop.group(1)
        m_img = re.search(r'hasImage:\s*true', obj_str)
        band['hasImage'] = True if m_img else False
        if 'name' in band and 'slug' in band:
            bands_data.append(band)

# Styles
header_fill = PatternFill(start_color='00205B', end_color='00205B', fill_type='solid') # Brand Navy
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')

gold_fill = PatternFill(start_color='FFB81C', end_color='FFB81C', fill_type='solid') # Brand Gold
gold_font = Font(name='Segoe UI', size=11, bold=True, color='00205B')

pending_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid') # Soft Gold
approved_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid') # Soft Green
rejected_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid') # Soft Red

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

wb = openpyxl.Workbook()
wb.remove(wb.active)

def create_tab(sheet_name, headers, rows_data, is_gold_header=False):
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    fill = gold_fill if is_gold_header else header_fill
    font = gold_font if is_gold_header else header_font
    
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    for row_idx, r in enumerate(rows_data, start=2):
        ws.append(r)
        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_num)
            c.font = Font(name='Segoe UI', size=9.5)
            c.border = thin_border
            if col_num in [1, 3, 4, 7, 13]:
                c.alignment = Alignment(horizontal='center')
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 52)
        
    return ws

# Data Validation Objects
dv_status = DataValidation(type="list", formula1='"Pending,Approved,Rejected,Archived"', allow_blank=True)
dv_category = DataValidation(type="list", formula1='"music,comedy,festival,dance,theatre,community"', allow_blank=True)
dv_cities = DataValidation(type="list", formula1='"Anchorage,Fairbanks,Juneau,Homer,Palmer,Wasilla,Seward,Girdwood,Talkeetna,Kenai,Soldotna,Ketchikan,Sitka,Kodiak,Valdez,Barrow,Bethel,Cordova,Haines,Skagway,Sterling,McCarthy,North Pole"', allow_blank=True)
dv_type = DataValidation(type="list", formula1='"Indoor,Outdoor/Festival,Hybrid"', allow_blank=True)
dv_claim_status = DataValidation(type="list", formula1='"Unclaimed,Pending Review,Verified Owner,Flagged"', allow_blank=True)
dv_account_status = DataValidation(type="list", formula1='"Active,Inactive,Closed"', allow_blank=True)
dv_artist_status = DataValidation(type="list", formula1='"Verified Artist,Unverified,Featured Legend,Touring Act"', allow_blank=True)
dv_feed_engine = DataValidation(type="list", formula1='"Facebook API,Web Scraper,iCal Feed,CSV Import"', allow_blank=True)
dv_menu_loc = DataValidation(type="list", formula1='"Header Desktop & Mobile,Bottom Mobile Pill,Global Category Filter"', allow_blank=True)

# 0. Executive Dashboard Tab
ws_dash = wb.create_sheet(title="0. Executive Dashboard", index=0)
ws_dash.views.sheetView[0].showGridLines = True

ws_dash['A1'] = "AK CONCERTS — EXECUTIVE CMS & CRM DASHBOARD"
ws_dash['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='00205B')
ws_dash['A2'] = "Master Database Control Panel & Website Publication Center for Cody"
ws_dash['A2'].font = Font(name='Segoe UI', size=11, color='64748B')

kpis = [
    ("Total Master Events", f"=COUNTA('2. Events_Master'!B2:B6000)", "A4", "B4"),
    ("Pending Web Submissions", f"=COUNTIF('1. Pending_Submissions'!M2:M1000, \"Pending\")", "D4", "E4"),
    ("Verified Venues", f"=COUNTA('3. Venues_CRM'!B2:B500)", "A6", "B6"),
    ("Local Artists Roster", f"=COUNTA('4. Bands_Artists_CRM'!B2:B500)", "D6", "E6")
]

for label, formula, cell_lbl, cell_val in kpis:
    ws_dash[cell_lbl] = label
    ws_dash[cell_lbl].font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    ws_dash[cell_lbl].fill = header_fill
    ws_dash[cell_lbl].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_dash[cell_val] = formula
    ws_dash[cell_val].font = Font(name='Segoe UI', size=16, bold=True, color='00205B')
    ws_dash[cell_val].fill = gold_fill
    ws_dash[cell_val].alignment = Alignment(horizontal='center', vertical='center')

ws_dash.column_dimensions['A'].width = 24
ws_dash.column_dimensions['B'].width = 16
ws_dash.column_dimensions['D'].width = 24
ws_dash.column_dimensions['E'].width = 16

ws_dash['A9'] = "⚡ CODY'S QUICK OPERATIONAL GUIDE"
ws_dash['A9'].font = Font(name='Segoe UI', size=12, bold=True, color='00205B')

guide_lines = [
    "1. Review New Submissions: Go to '1. Pending_Submissions' tab to review web submissions from /submit.",
    "2. Change Status to Approve: Click the Status dropdown chip in Column M and change from 'Pending' to 'Approved'.",
    "3. Automated Site Rebuild: Once approved, run '⚡ AK Concerts Controls ➔ Trigger Live Website Build' in the menu bar.",
    "4. Master Event Database: All approved events sync into '2. Events_Master' and deploy to 3,500+ static pages.",
    "5. Managing Venues & Artists: Update venue contacts in '3. Venues_CRM' and artist video IDs in '4. Bands_Artists_CRM'."
]

for idx, g_text in enumerate(guide_lines, start=10):
    ws_dash[f'A{idx}'] = g_text
    ws_dash[f'A{idx}'].font = Font(name='Segoe UI', size=10, color='1E293B')

# 1. Pending_Submissions Tab
ws_pending = create_tab(
    "1. Pending_Submissions",
    ["Timestamp", "Title", "Date", "Time", "City", "Venue", "Category", "Cost", "TicketUrl", "Email", "SubmitterName", "Notes", "Status"],
    [
        ["2026-08-12 22:45:00", "Blackwater Railroad Live at Williwaw", "2026-08-22", "8:00 PM", "Anchorage", "Williwaw Social", "music", "$15", "https://akconcerts.com", "promoter@blackwaterrr.com", "Cody Submitter", "Web Submission via /submit page", "Pending"],
        ["2026-08-12 23:10:00", "Stand-Up Comedy Night w/ Dana Gould", "2026-08-23", "7:30 PM", "Fairbanks", "Pioneer Park Theater", "comedy", "$25", "https://pioneerpark.org", "boxoffice@pioneerpark.org", "Sarah Admin", "Venue Claim Test Submission", "Pending"]
    ],
    is_gold_header=True
)
ws_pending.add_data_validation(dv_status)
dv_status.add("M2:M1000")
ws_pending.add_data_validation(dv_category)
dv_category.add("G2:G1000")
ws_pending.add_data_validation(dv_cities)
dv_cities.add("E2:E1000")

rule_pending = CellIsRule(operator='equal', formula=['"Pending"'], fill=pending_fill, font=Font(color='92400E', bold=True))
rule_approved = CellIsRule(operator='equal', formula=['"Approved"'], fill=approved_fill, font=Font(color='166534', bold=True))
rule_rejected = CellIsRule(operator='equal', formula=['"Rejected"'], fill=rejected_fill, font=Font(color='991B1B', bold=True))

ws_pending.conditional_formatting.add("M2:M1000", rule_pending)
ws_pending.conditional_formatting.add("M2:M1000", rule_approved)
ws_pending.conditional_formatting.add("M2:M1000", rule_rejected)

# 2. Events_Master Tab
master_event_rows = []
for e in events_data:
    master_event_rows.append([
        "2026-08-12 12:00:00",
        e.get("title", ""),
        e.get("date", ""),
        e.get("time", ""),
        e.get("city", ""),
        e.get("venue", ""),
        e.get("category", "music"),
        e.get("cost", "") or e.get("price", ""),
        e.get("ticketUrl", ""),
        "system@akconcerts.com",
        "AK Concerts Auto Import",
        "Master Event Database",
        "Approved"
    ])
ws_events = create_tab("2. Events_Master", ["Timestamp", "Title", "Date", "Time", "City", "Venue", "Category", "Cost", "TicketUrl", "Email", "SubmitterName", "Notes", "Status"], master_event_rows)

dv_status_master = DataValidation(type="list", formula1='"Pending,Approved,Rejected,Archived"', allow_blank=True)
ws_events.add_data_validation(dv_status_master)
dv_status_master.add(f"M2:M{len(master_event_rows)+10}")

dv_category_master = DataValidation(type="list", formula1='"music,comedy,festival,dance,theatre,community"', allow_blank=True)
ws_events.add_data_validation(dv_category_master)
dv_category_master.add(f"G2:G{len(master_event_rows)+10}")

dv_cities_master = DataValidation(type="list", formula1='"Anchorage,Fairbanks,Juneau,Homer,Palmer,Wasilla,Seward,Girdwood,Talkeetna,Kenai,Soldotna,Ketchikan,Sitka,Kodiak,Valdez,Barrow,Bethel,Cordova,Haines,Skagway,Sterling,McCarthy,North Pole"', allow_blank=True)
ws_events.add_data_validation(dv_cities_master)
dv_cities_master.add(f"E2:E{len(master_event_rows)+10}")

ws_events.conditional_formatting.add(f"M2:M{len(master_event_rows)+10}", rule_approved)
ws_events.conditional_formatting.add(f"M2:M{len(master_event_rows)+10}", rule_pending)
ws_events.conditional_formatting.add(f"M2:M{len(master_event_rows)+10}", rule_rejected)

# 3. Venues_CRM Tab (ALL 183+ VENUES)
venue_rows = []
for v in venues_data:
    venue_rows.append([
        v.get("id", ""),
        v.get("name", ""),
        v.get("city", ""),
        v.get("address", "") or f"{v.get('name')}, {v.get('city')}, AK",
        v.get("capacity", "") or "N/A",
        "Indoor" if v.get("indoor") else "Outdoor/Festival",
        v.get("website", ""),
        v.get("facebookUrl", ""),
        v.get("lat", "") or "",
        v.get("lng", "") or "",
        "Verified Owner" if v.get("website") else "Unclaimed",
        "Active"
    ])
ws_venues = create_tab("3. Venues_CRM", ["Venue ID", "Venue Name", "City", "Address", "Capacity", "Type", "Website", "Facebook Page", "Latitude", "Longitude", "Claim Status", "Account Status"], venue_rows)

ws_venues.add_data_validation(dv_type)
dv_type.add(f"F2:F{len(venue_rows)+10}")

ws_venues.add_data_validation(dv_claim_status)
dv_claim_status.add(f"K2:K{len(venue_rows)+10}")

ws_venues.add_data_validation(dv_account_status)
dv_account_status.add(f"L2:L{len(venue_rows)+10}")

# 4. Bands_Artists_CRM Tab (ALL 54+ BANDS)
bands_rows = []
for b in bands_data:
    bands_rows.append([
        b.get("slug", ""),
        b.get("name", ""),
        "Local Artist",
        "Alaska",
        b.get("description", ""),
        b.get("facebookUrl", ""),
        b.get("youtubeId", ""),
        "Yes" if b.get("hasImage") else "No",
        "Verified Artist"
    ])
ws_bands = create_tab("4. Bands_Artists_CRM", ["Band Slug", "Band Name", "Genre", "Home City", "Bio Description", "Facebook URL", "YouTube Video ID", "Has Image", "Profile Status"], bands_rows)

ws_bands.add_data_validation(dv_artist_status)
dv_artist_status.add(f"I2:I{len(bands_rows)+10}")

# 5. Cities_Regions Tab
city_rows = [
    ["Anchorage", "Greater Anchorage Borough", "Southcentral", 61.2181, -149.9003, "Active Core"],
    ["Fairbanks", "Fairbanks North Star Borough", "Interior", 64.8378, -147.7164, "Active Core"],
    ["Juneau", "City & Borough of Juneau", "Southeast", 58.3019, -134.4197, "Active Core"],
    ["Homer", "Kenai Peninsula Borough", "Kenai Peninsula", 59.6425, -151.5483, "Active Regional"],
    ["Palmer", "Matanuska-Susitna Borough", "Mat-Su Valley", 61.5997, -149.1100, "Active Regional"],
    ["Wasilla", "Matanuska-Susitna Borough", "Mat-Su Valley", 61.5814, -149.4394, "Active Regional"],
    ["Girdwood", "Municipality of Anchorage", "Southcentral", 60.9422, -149.1678, "Active Regional"],
    ["Seward", "Kenai Peninsula Borough", "Kenai Peninsula", 60.1042, -149.4422, "Active Regional"],
    ["Talkeetna", "Matanuska-Susitna Borough", "Mat-Su Valley", 62.3208, -150.1066, "Active Regional"],
    ["Ketchikan", "Ketchikan Gateway Borough", "Southeast", 55.3422, -131.6461, "Active Regional"],
    ["Sitka", "City and Borough of Sitka", "Southeast", 57.0531, -135.3300, "Active Regional"]
]
ws_cities = create_tab("5. Cities_Regions", ["City Name", "Borough / Region", "Territory Zone", "Latitude", "Longitude", "Coverage Level"], city_rows)

# 6. Promoters_Partners Tab
promoters_rows = [
    ["PROMO-01", "Showdown Alaska", "Major Touring & Festivals", "Anchorage", "info@showdownalaska.com", "https://showdownalaska.com", "Active Key Partner"],
    ["PROMO-02", "Whale Bell Media", "Independent Music & Concerts", "Homer", "events@whalebell.com", "https://whalebell.com", "Active Partner"],
    ["PROMO-03", "Alaska Center for the Performing Arts", "Broadway & Performing Arts", "Anchorage", "boxoffice@alaskapac.org", "https://alaskapac.org", "Active Venue Partner"],
    ["PROMO-04", "Creekbend Company", "Outdoor Summer Concert Series", "Hope", "info@creekbendco.com", "https://creekbendco.com", "Active Festival Partner"]
]
ws_promoters = create_tab("6. Promoters_Partners", ["Promoter ID", "Organization Name", "Specialty Type", "Base City", "Contact Email", "Website", "Partner Tier"], promoters_rows)

# 7. Scraper_Sources Tab
scrapers_rows = [
    ["SCRAPE-01", "Williwaw Social Facebook Events", "Facebook API", "Daily (02:00 AKDT)", "Active", "2026-08-12 02:00"],
    ["SCRAPE-02", "Koot's Nightclub Schedule", "Web Scraper", "Daily (02:30 AKDT)", "Active", "2026-08-12 02:30"],
    ["SCRAPE-03", "49th State Brewing Co Feed", "iCal / Web Scraper", "Daily (03:00 AKDT)", "Active", "2026-08-12 03:00"],
    ["SCRAPE-04", "Crystal Saloon Juneau Scraper", "Web Scraper", "Daily (03:30 AKDT)", "Active", "2026-08-12 03:30"],
    ["SCRAPE-05", "Palmer Alehouse Calendar", "Web Scraper", "Daily (04:00 AKDT)", "Active", "2026-08-12 04:00"],
    ["SCRAPE-06", "Salmonfest Alaska Master Import", "CSV Import", "Weekly", "Active", "2026-08-10 12:00"]
]
ws_scrapers = create_tab("7. Scraper_Sources", ["Scraper ID", "Source Feed Name", "Feed Engine", "Scrape Frequency", "Status", "Last Run Timestamp"], scrapers_rows)

ws_scrapers.add_data_validation(dv_feed_engine)
dv_feed_engine.add(f"C2:C{len(scrapers_rows)+10}")

# 8. App_Settings_Menus Tab
settings_rows = [
    ["HEADER_NAV_1", "Playing Soon", "/", "Header Desktop & Mobile", "Enabled"],
    ["HEADER_NAV_2", "Past Events", "/past", "Header Desktop & Mobile", "Enabled"],
    ["HEADER_NAV_3", "Submit Event", "/submit", "Header Desktop & Mobile", "Enabled"],
    ["HEADER_NAV_4", "Bands", "/bands", "Header Desktop & Mobile", "Enabled"],
    ["HEADER_NAV_5", "Venues", "/venues", "Header Desktop & Mobile", "Enabled"],
    ["HEADER_NAV_6", "Newsletter", "/subscribe", "Header Desktop & Mobile", "Enabled"],
    ["FLOOR_NAV_1", "Shows", "/", "Bottom Mobile Pill", "Enabled"],
    ["FLOOR_NAV_2", "Bands", "/bands", "Bottom Mobile Pill", "Enabled"],
    ["FLOOR_NAV_3", "Venues", "/venues", "Bottom Mobile Pill", "Enabled"],
    ["FLOOR_NAV_4", "Past", "/past", "Bottom Mobile Pill", "Enabled"],
    ["FLOOR_NAV_5", "Submit", "/submit", "Bottom Mobile Pill", "Enabled"],
    ["CATEGORY_1", "Live Music", "music", "Global Category Filter", "Enabled"],
    ["CATEGORY_2", "Comedy", "comedy", "Global Category Filter", "Enabled"],
    ["CATEGORY_3", "Festivals", "festival", "Global Category Filter", "Enabled"],
    ["CATEGORY_4", "Dance", "dance", "Global Category Filter", "Enabled"],
    ["CATEGORY_5", "Theatre & Arts", "theatre", "Global Category Filter", "Enabled"],
    ["CATEGORY_6", "Community", "community", "Global Category Filter", "Enabled"]
]
ws_settings = create_tab("8. App_Settings_Menus", ["Setting Key", "Label Name", "Target Route / Slug", "Menu Location", "Status"], settings_rows)

ws_settings.add_data_validation(dv_menu_loc)
dv_menu_loc.add(f"D2:D{len(settings_rows)+10}")

# Save Workbook
for path in [output_xlsx_path, artifact_xlsx_path]:
    wb.save(path)

print(f"Successfully updated Master CMS/CRM Excel Workbook with ALL {len(bands_data)} Bands and ALL {len(venue_rows)} Venues: {output_xlsx_path}")
